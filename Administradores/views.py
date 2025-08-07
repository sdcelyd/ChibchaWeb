from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.views import LoginView
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.urls import reverse_lazy
from django.db.models import Sum, Count, Q, Value, CharField
from django.utils import timezone
from datetime import datetime, timedelta
from django.core.paginator import Paginator
from django.http import JsonResponse
import json

from Administradores.forms import EditarUsuarioForm
from ChibchaWeb.decorators import administrador_required, admin_permission_required
from Clientes.forms import RegistroClienteForm
from .models import Administrador
from Clientes.models import Cliente
from Empleados.models import Empleado  
from Pagos.models import Pago, PagoDistribuidor
#from .forms import CrearUsuarioForm, EditarUsuarioForm
import csv
import io
import pandas as pd
from django.http import HttpResponse
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO

class AdministradorLoginView(LoginView):
    template_name = 'administradores/login.html'
    redirect_authenticated_user = True
    
    def get_success_url(self):
        return reverse_lazy('administradores:dashboard')
    
    def form_valid(self, form):
        """Verificar que el usuario sea administrador antes de autenticar"""
        username = form.cleaned_data.get('username')
        password = form.cleaned_data.get('password')
        
        user = authenticate(self.request, username=username, password=password)
        if user is not None:
            try:
                administrador = Administrador.objects.get(user=user)
                if not administrador.activo:
                    messages.error(self.request, 'Tu cuenta de administrador está desactivada.')
                    return self.form_invalid(form)
                    
                # Actualizar último acceso
                administrador.ultimo_acceso = timezone.now()
                administrador.save()
                
                login(self.request, user)
                messages.success(self.request, f'Bienvenido {user.username}. Redirigiendo al panel de administración.')
                return redirect('administradores:dashboard')
            except Administrador.DoesNotExist:
                messages.error(self.request, 'No tienes permisos de administrador para acceder.')
                return self.form_invalid(form)
        else:
            messages.error(self.request, 'Usuario o contraseña incorrectos.')
            return self.form_invalid(form)


@administrador_required
def dashboard(request):
    """Dashboard principal del administrador"""
    # Estadísticas básicas
    total_clientes = Cliente.objects.count()
    total_empleados = Empleado.objects.count()
    total_pagos = Pago.objects.count()
    ingresos_totales = Pago.objects.aggregate(total=Sum('monto'))['total'] or 0
    
    # Ingresos del último mes
    ultimo_mes = timezone.now() - timedelta(days=30)
    ingresos_ultimo_mes = Pago.objects.filter(
        fecha__gte=ultimo_mes
    ).aggregate(total=Sum('monto'))['total'] or 0
    
    # Clientes con suscripción activa
    clientes_activos = Cliente.objects.filter(tiene_suscripcion=True).count()
    
    context = {
        'total_clientes': total_clientes,
        'total_empleados': total_empleados,
        'total_pagos': total_pagos,
        'ingresos_totales': ingresos_totales,
        'ingresos_ultimo_mes': ingresos_ultimo_mes,
        'clientes_activos': clientes_activos,
    }
    
    return render(request, 'administradores/dashboard.html', context)


@admin_permission_required('puede_gestionar_usuarios')
def gestionar_usuarios(request):
    """Vista para gestionar todos los usuarios"""
    # Filtros
    tipo_usuario = request.GET.get('tipo', 'all')
    busqueda = request.GET.get('q', '')
    
    # Obtener usuarios según filtros
    usuarios = User.objects.all()
    
    if busqueda:
        usuarios = usuarios.filter(
            Q(username__icontains=busqueda) |
            Q(email__icontains=busqueda) |
            Q(first_name__icontains=busqueda) |
            Q(last_name__icontains=busqueda)
        )
    
    # Preparar datos con información adicional
    usuarios_data = []
    for user in usuarios:
        user_info = {
            'user': user,
            'tipo': 'Usuario base',
            'activo': user.is_active,
            'additional_info': None
        }
        
        # Verificar tipo de usuario
        if hasattr(user, 'cliente'):
            user_info['tipo'] = 'Cliente'
            user_info['additional_info'] = user.cliente
            user_info['activo'] = user.cliente.activo if hasattr(user.cliente, 'activo') else user.is_active
        elif hasattr(user, 'empleado'):
            user_info['tipo'] = 'Empleado'
            user_info['additional_info'] = user.empleado
            user_info['activo'] = user.empleado.activo
        elif hasattr(user, 'administrador'):
            user_info['tipo'] = 'Administrador'
            user_info['additional_info'] = user.administrador
            user_info['activo'] = user.administrador.activo
        
        # Filtrar por tipo si se especifica
        if tipo_usuario != 'all' and user_info['tipo'].lower() != tipo_usuario.lower():
            continue
            
        usuarios_data.append(user_info)
    
    # Paginación
    paginator = Paginator(usuarios_data, 20)  # 20 usuarios por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'tipo_usuario': tipo_usuario,
        'busqueda': busqueda,
    }
    
    return render(request, 'administradores/gestionar_usuarios.html', context)


@admin_permission_required('puede_gestionar_usuarios')
def crear_usuario(request):
    """Vista para crear nuevos usuarios"""
    from .forms import CrearUsuarioForm  # Importar el formulario correcto
    
    if request.method == 'POST':
        form = CrearUsuarioForm(request.POST)
        if form.is_valid():
            try:
                # Crear usuario base
                user = User.objects.create_user(
                    username=form.cleaned_data['username'],
                    email=form.cleaned_data['email'],
                    password=form.cleaned_data['password1'],
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name']
                )
                
                # Crear perfil específico según el tipo
                tipo_usuario = form.cleaned_data['tipo_usuario']
                telefono = form.cleaned_data.get('telefono', '')
                
                if tipo_usuario == 'cliente':
                    Cliente.objects.create(
                        user=user,
                        telefono=telefono,
                        activo=True
                    )
                elif tipo_usuario == 'empleado':
                    # Aquí es donde se crea el empleado con el rol seleccionado
                    Empleado.objects.create(
                        user=user,
                        telefono=telefono,
                        rol=form.cleaned_data['rol_empleado'],
                        activo=True,
                        nivel=1  # Nivel por defecto
                    )
                elif tipo_usuario == 'administrador':
                    Administrador.objects.create(
                        user=user,
                        activo=True
                    )
                
                messages.success(request, f'Usuario {user.username} creado exitosamente como {tipo_usuario}.')
                return redirect('administradores:gestionar_usuarios')
                
            except Exception as e:
                messages.error(request, f'Error al crear usuario: {str(e)}')
                # Si hay error, eliminar el usuario creado para evitar inconsistencias
                if 'user' in locals():
                    user.delete()
    else:
        form = CrearUsuarioForm()
    
    return render(request, 'administradores/crear_usuario.html', {'form': form})


@admin_permission_required('puede_gestionar_usuarios')
def editar_usuario(request, user_id):
    """Vista para editar usuarios existentes con capacidad de cambiar tipos"""
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        form = EditarUsuarioForm(request.POST, instance=user)
        if form.is_valid():
            try:
                # Obtener valores antes de guardar
                tipo_usuario = form.cleaned_data.get('tipo_usuario')
                activo = form.cleaned_data.get('activo', True)
                is_active = form.cleaned_data.get('is_active', True)
                
                # Guardar el usuario base
                user = form.save(commit=False)
                user.is_active = is_active
                user.save()
                
                # Manejar perfiles específicos manualmente para asegurar sincronización
                if tipo_usuario == 'cliente':
                    # Eliminar otros perfiles si existen
                    if hasattr(user, 'empleado'):
                        user.empleado.delete()
                    if hasattr(user, 'administrador'):
                        user.administrador.delete()
                    
                    # Crear o actualizar cliente
                    cliente, created = Cliente.objects.get_or_create(
                        user=user,
                        defaults={
                            'telefono': form.cleaned_data.get('telefono', ''),
                            'activo': activo
                        }
                    )
                    if not created:
                        cliente.telefono = form.cleaned_data.get('telefono', '')
                        cliente.activo = activo
                        cliente.save()
                        
                elif tipo_usuario == 'empleado':
                    # Eliminar otros perfiles si existen
                    if hasattr(user, 'cliente'):
                        user.cliente.delete()
                    if hasattr(user, 'administrador'):
                        user.administrador.delete()
                    
                    # Crear o actualizar empleado
                    empleado, created = Empleado.objects.get_or_create(
                        user=user,
                        defaults={
                            'telefono': form.cleaned_data.get('telefono', ''),
                            'rol': form.cleaned_data.get('rol_empleado', 'agente'),
                            'nivel': form.cleaned_data.get('nivel_empleado', 1),
                            'activo': activo
                        }
                    )
                    if not created:
                        empleado.telefono = form.cleaned_data.get('telefono', '')
                        empleado.rol = form.cleaned_data.get('rol_empleado', 'agente')
                        empleado.nivel = form.cleaned_data.get('nivel_empleado', 1)
                        empleado.activo = activo
                        empleado.save()
                        
                elif tipo_usuario == 'administrador':
                    # Eliminar otros perfiles si existen
                    if hasattr(user, 'cliente'):
                        user.cliente.delete()
                    if hasattr(user, 'empleado'):
                        user.empleado.delete()
                    
                    # Crear o actualizar administrador
                    administrador, created = Administrador.objects.get_or_create(
                        user=user,
                        defaults={'activo': activo}
                    )
                    if not created:
                        administrador.activo = activo
                        administrador.save()
                
                messages.success(
                    request, 
                    f'Usuario {user.username} actualizado exitosamente.'
                )
                return redirect('administradores:gestionar_usuarios')
                
            except Exception as e:
                messages.error(request, f'Error al actualizar usuario: {str(e)}')
    else:
        form = EditarUsuarioForm(instance=user)
    
    # Información adicional para el contexto
    tipo_actual = 'Usuario base'
    info_adicional = None
    
    if hasattr(user, 'cliente'):
        tipo_actual = 'Cliente'
        info_adicional = user.cliente
    elif hasattr(user, 'empleado'):
        tipo_actual = 'Empleado'
        info_adicional = user.empleado
    elif hasattr(user, 'administrador'):
        tipo_actual = 'Administrador'
        info_adicional = user.administrador
    
    context = {
        'form': form,
        'user_obj': user,
        'tipo_actual': tipo_actual,
        'info_adicional': info_adicional,
    }
    
    return render(request, 'administradores/editar_usuario.html', context)


@admin_permission_required('puede_gestionar_usuarios')
def eliminar_usuario(request, user_id):
    """Vista para eliminar usuarios"""
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        username = user.username
        user.delete()
        messages.success(request, f'Usuario {username} eliminado exitosamente.')
        return redirect('administradores:gestionar_usuarios')
    
    return render(request, 'administradores/confirmar_eliminacion.html', {'user_obj': user})


@admin_permission_required('puede_ver_estadisticas')
def estadisticas_pagos(request):
    """Vista para mostrar estadísticas de pagos de hosting y distribuidores"""
    # Parámetros de filtro
    periodo = request.GET.get('periodo', '30')  # días
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    # Construir filtros de fecha
    filtros = Q()
    
    if fecha_inicio and fecha_fin:
        try:
            inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            filtros &= Q(fecha__date__gte=inicio, fecha__date__lte=fin)
        except ValueError:
            messages.error(request, 'Formato de fecha inválido.')
    else:
        # Usar período predefinido
        dias = int(periodo)
        fecha_limite = timezone.now() - timedelta(days=dias)
        filtros &= Q(fecha__gte=fecha_limite)
    
    # Obtener pagos de hosting filtrados
    pagos_hosting = Pago.objects.filter(filtros).exclude(pagodistribuidor__isnull=False)
    
    # Obtener pagos de distribuidores filtrados
    pagos_distribuidores = PagoDistribuidor.objects.filter(filtros)
    
    # Estadísticas de hosting
    total_ingresos_hosting = pagos_hosting.aggregate(total=Sum('monto'))['total'] or 0
    total_transacciones_hosting = pagos_hosting.count()
    
    # Estadísticas de distribuidores
    total_ingresos_distribuidores = pagos_distribuidores.aggregate(total=Sum('monto'))['total'] or 0
    total_transacciones_distribuidores = pagos_distribuidores.count()
    total_paginas_vendidas = pagos_distribuidores.aggregate(total=Sum('cantidad_paginas'))['total'] or 0
    
    # Estadísticas generales combinadas
    total_ingresos = total_ingresos_hosting + total_ingresos_distribuidores
    total_transacciones = total_transacciones_hosting + total_transacciones_distribuidores
    promedio_transaccion = total_ingresos / total_transacciones if total_transacciones > 0 else 0
    
    # Ingresos por día de hosting (últimos 30 días para gráfico)
    ultimos_30_dias = timezone.now() - timedelta(days=30)
    ingresos_diarios_hosting = (
        pagos_hosting.filter(fecha__gte=ultimos_30_dias)
        .extra({'fecha_dia': "date(fecha)"})
        .values('fecha_dia')
        .annotate(total=Sum('monto'), cantidad=Count('pagoId'))
        .order_by('fecha_dia')
    )
    
    # Ingresos por día de distribuidores (últimos 30 días para gráfico)
    ingresos_diarios_distribuidores = (
        PagoDistribuidor.objects.filter(fecha__gte=ultimos_30_dias)
        .extra({'fecha_dia': "date(fecha)"})
        .values('fecha_dia')
        .annotate(total=Sum('monto'), cantidad=Count('pagoId'))
        .order_by('fecha_dia')
    )
    
    # Ingresos por plan de hosting
    ingresos_por_plan = (
        pagos_hosting.values('cliente__plan')
        .annotate(total=Sum('monto'), cantidad=Count('pagoId'))
        .order_by('-total')
    )
    
    # Top clientes por ingresos (ambos tipos)
    
    # Pagos de hosting
    hosting_clientes = pagos_hosting.values('cliente__user__username', 'cliente__user__email').annotate(
        total=Sum('monto'), 
        cantidad=Count('pagoId'),
        tipo=Value('Hosting', output_field=CharField())
    )
    
    # Pagos de distribuidores
    distribuidores_clientes = pagos_distribuidores.values('cliente__user__username', 'cliente__user__email').annotate(
        total=Sum('monto'), 
        cantidad=Count('pagoId'),
        tipo=Value('Distribuidor', output_field=CharField())
    )
    
    # Top distribuidores por páginas compradas
    top_distribuidores = (
        pagos_distribuidores.values('cliente__user__username', 'cliente__user__email')
        .annotate(
            total=Sum('monto'), 
            cantidad=Count('pagoId'),
            paginas=Sum('cantidad_paginas')
        )
        .order_by('-paginas')[:10]
    )
    
    context = {
        # Estadísticas generales
        'total_ingresos': total_ingresos,
        'total_transacciones': total_transacciones,
        'promedio_transaccion': promedio_transaccion,
        
        # Estadísticas por tipo
        'total_ingresos_hosting': total_ingresos_hosting,
        'total_transacciones_hosting': total_transacciones_hosting,
        'total_ingresos_distribuidores': total_ingresos_distribuidores,
        'total_transacciones_distribuidores': total_transacciones_distribuidores,
        'total_paginas_vendidas': total_paginas_vendidas,
        
        # Datos para gráficos
        'ingresos_diarios_hosting': list(ingresos_diarios_hosting),
        'ingresos_diarios_distribuidores': list(ingresos_diarios_distribuidores),
        'ingresos_por_plan': list(ingresos_por_plan),
        'hosting_clientes': list(hosting_clientes)[:10],
        'distribuidores_clientes': list(distribuidores_clientes)[:10],
        'top_distribuidores': list(top_distribuidores),
        
        # Filtros
        'periodo': periodo,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    }
    
    return render(request, 'administradores/estadisticas_pagos.html', context)


@admin_permission_required('puede_ver_estadisticas')
def estadisticas_usuarios(request):
    """Vista para mostrar estadísticas de usuarios"""
    # Estadísticas de clientes
    total_clientes = Cliente.objects.count()
    clientes_con_suscripcion = Cliente.objects.filter(tiene_suscripcion=True).count()
    
    porcentaje_suscripcion = (clientes_con_suscripcion * 100 / total_clientes) if total_clientes > 0 else 0

    # Estadísticas de empleados
    total_empleados = Empleado.objects.count()
    empleados_activos = Empleado.objects.filter(activo=True).count()
    empleados_por_rol = (
        Empleado.objects.values('rol')
        .annotate(cantidad=Count('id'))
        .order_by('rol')
    )
    
    # Registros recientes (últimos 30 días)
    ultimos_30_dias = timezone.now() - timedelta(days=30)
    clientes_recientes = Cliente.objects.filter(
        user__date_joined__gte=ultimos_30_dias
    ).count()
    
    # Distribución por planes
    planes_distribucion = (
        Cliente.objects.filter(tiene_suscripcion=True)
        .values('plan')
        .annotate(cantidad=Count('id'))
        .order_by('plan')
    )
    
    context = {
        'total_clientes': total_clientes,
        'clientes_con_suscripcion': clientes_con_suscripcion,
        'porcentaje_suscripcion': round(porcentaje_suscripcion, 1),
        'total_empleados': total_empleados,
        'empleados_activos': empleados_activos,
        'empleados_por_rol': list(empleados_por_rol),
        'clientes_recientes': clientes_recientes,
        'planes_distribucion': list(planes_distribucion),
    }
    
    return render(request, 'administradores/estadisticas_usuarios.html', context)


@admin_permission_required('puede_ver_estadisticas')
def api_datos_grafico(request):
    """API para obtener datos de gráficos via AJAX"""
    tipo = request.GET.get('tipo')
    
    if tipo == 'ingresos_diarios':
        dias = int(request.GET.get('dias', 30))
        fecha_limite = timezone.now() - timedelta(days=dias)
        
        datos = (
            Pago.objects.filter(fecha__gte=fecha_limite)
            .extra({'fecha_dia': "date(fecha)"})
            .values('fecha_dia')
            .annotate(total=Sum('monto'))
            .order_by('fecha_dia')
        )
        
        return JsonResponse({
            'labels': [item['fecha_dia'].strftime('%Y-%m-%d') for item in datos],
            'data': [float(item['total']) for item in datos]
        })
    
    elif tipo == 'planes_distribucion':
        datos = (
            Cliente.objects.filter(tiene_suscripcion=True)
            .values('plan')
            .annotate(cantidad=Count('clienteId'))
            .order_by('plan')
        )
        
        return JsonResponse({
            'labels': [item['plan'] or 'Sin Plan' for item in datos],
            'data': [item['cantidad'] for item in datos]
        })
    
    return JsonResponse({'error': 'Tipo de gráfico no válido'}, status=400)

@admin_permission_required('puede_ver_estadisticas')
def exportar_pagos_csv(request):
    """Exportar datos de pagos a CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="pagos.csv"'
    response.write('\ufeff')  # BOM para UTF-8
    
    writer = csv.writer(response)
    
    # Encabezados
    writer.writerow([
        'ID Pago', 'Cliente', 'Email Cliente', 'Monto', 
        'Fecha', 'Estado', 'Método Pago', 'Plan'
    ])
    
    # Obtener pagos con información del cliente
    pagos = Pago.objects.all().select_related(
        'cliente__user'
    ).order_by('-fecha')
    
    for pago in pagos:
        writer.writerow([
            pago.pagoId,
            f"{pago.cliente.user.first_name} {pago.cliente.user.last_name}".strip() or pago.cliente.user.username,
            pago.cliente.user.email,
            pago.monto,
            pago.fecha.strftime('%Y-%m-%d %H:%M'),
            getattr(pago, 'estado', 'Completado'),
            getattr(pago, 'metodo_pago', 'N/A'),
            getattr(pago.cliente, 'plan', 'N/A')
        ])
    
    return response


@admin_permission_required('puede_ver_estadisticas')
def exportar_pagos_excel(request):
    """Exportar datos de pagos a Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'ID Pago', 'Cliente', 'Email Cliente', 'Monto', 
        'Fecha', 'Estado', 'Método Pago', 'Plan'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    pagos = Pago.objects.all().select_related(
        'cliente__user'
    ).order_by('-fecha')
    
    for row, pago in enumerate(pagos, 2):
        data_row = [
            pago.pagoId,
            f"{pago.cliente.user.first_name} {pago.cliente.user.last_name}".strip() or pago.cliente.user.username,
            pago.cliente.user.email,
            float(pago.monto),
            pago.fecha.strftime('%Y-%m-%d %H:%M'),
            getattr(pago, 'estado', 'Completado'),
            getattr(pago, 'metodo_pago', 'N/A'),
            getattr(pago.cliente, 'plan', 'N/A')
        ]
        
        for col, value in enumerate(data_row, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pagos.xlsx"'
    
    wb.save(response)
    return response


@admin_permission_required('puede_ver_estadisticas')
def exportar_pagos_pdf(request):
    """Exportar datos de pagos a PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72,
                          topMargin=72, bottomMargin=18)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    story = []
    
    # Título
    title = Paragraph("Reporte de Pagos", title_style)
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Fecha de generación
    fecha = timezone.now().strftime('%d/%m/%Y %H:%M')
    fecha_p = Paragraph(f"<b>Fecha de generación:</b> {fecha}", styles['Normal'])
    story.append(fecha_p)
    story.append(Spacer(1, 12))
    
    # Obtener datos
    pagos = Pago.objects.all().select_related(
        'cliente__user'
    ).order_by('-fecha')[:100]  # Limitar a 100 registros más recientes
    
    # Estadísticas
    total_pagos = Pago.objects.count()
    total_ingresos = Pago.objects.aggregate(total=Sum('monto'))['total'] or 0
    
    stats_text = f"""
    <b>Estadísticas Generales:</b><br/>
    • Total de pagos: {total_pagos}<br/>
    • Ingresos totales: ${total_ingresos:,.2f}<br/>
    • Promedio por pago: ${(total_ingresos/total_pagos if total_pagos > 0 else 0):,.2f}<br/>
    <i>Mostrando los 100 pagos más recientes</i>
    """
    
    story.append(Paragraph(stats_text, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Crear tabla
    data = [['ID', 'Cliente', 'Monto', 'Fecha', 'Plan']]
    
    for pago in pagos:
        cliente_nombre = f"{pago.cliente.user.first_name} {pago.cliente.user.last_name}".strip() or pago.cliente.user.username
        data.append([
            str(pago.pagoId),
            cliente_nombre,
            f"${pago.monto:,.2f}",
            pago.fecha.strftime('%d/%m/%Y'),
            getattr(pago.cliente, 'plan', 'N/A')
        ])
    
    # Crear tabla con estilos
    table = Table(data, colWidths=[0.8*inch, 2*inch, 1*inch, 1.2*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(table)
    
    # Construir PDF
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pagos.pdf"'
    response.write(pdf)
    
    return response